
import openpyxl
import pandas as pd
import requests as req
from openpyxl.styles import Border, Font, PatternFill, Side

url = "http://139.59.36.44/api/v1/users/search/"


head_bd = Border(Side(style="thin", color="000000",),
                 Side(style="thin", color="000000",),
                 Side(style="thin", color="000000",),
                 Side(style="thin", color="000000",))
head_fn = Font(sz=12, bold=True)
head_fl = PatternFill(patternType='solid',
                      fgColor='E0E0E0')

mark_fl = PatternFill(patternType='solid',
                      fgColor='FC2C03')


all_df = pd.read_csv("xlxs_folder/installations/all.csv")


class DummyCell:
    def __init__(self, cell) -> None:
        self.cell = cell

    def __hash__(self) -> int:
        return hash(self.cell.value)

    def __eq__(self, __o: object) -> bool:
        return self.cell.value == __o

    def __str__(self) -> str:
        return f"<{self.__class__.__name__}: {self.cell.value}>"

    def __repr__(self) -> str:
        return self.__str__()

    def __lt__(self, __o) -> bool:
        return self.cell.value < __o

    def __gt__(self, __o) -> bool:
        return self.cell.value > __o


for col in all_df.columns:
    print(col)

    data = {"installations": list((i for i in all_df[col] if not pd.isna(i)))}

    res = req.post(url, json=data)

    if res:
        data = res.json()["data"]
        # data = [{'installation': 'ID501328B', 'plates': ['1-2-3-4']},
        #         {'installation': 'ID501328B',
        #             'plates': ['1-2-3-4', '1-5-3-4']},
        #         {'installation': 'ID501328B',
        #             'plates': ['1-2-3-6', '1-5-3-5']},
        #         {'installation': 'ID501328B',
        #             'plates': ['1-2-3-5', '1-5-4-4']},
        #         {'installation': 'ID502934A', 'plates': ['#NA']},
        #         {'installation': 'ID501328B',
        #             'plates': ['1-2-3-4', '1-5-3-4']},
        #         {'installation': 'ID501328B',
        #             'plates': ['2-2-3-6', '1-5-3-5']},
        #         ]

        # print(data)

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet2 = wb.create_sheet("Sheet2")

        cell = sheet["A1"]

        cell.value = "installation"
        cell.border = head_bd
        cell.fill = head_fl
        cell.font = head_fn

        cell = sheet["B1"]
        cell.value = "plates"
        cell.border = head_bd
        cell.fill = head_fl
        cell.font = head_fn

        cell = sheet2["A1"]
        cell.value = "plates"
        cell.border = head_bd
        cell.fill = head_fl
        cell.font = head_fn

        cache = {}

        plates = set()

        for idx, i in enumerate(data):

            sheet.cell(idx + 1 + 1, 1).value = i["installation"]

            v: str

            for i, v in enumerate(i["plates"]):

                c, r = idx + 1 + 1, i + 1 + 1
                cell = sheet.cell(c, r)
                cell.value = v
                s = v.split("-")
                _b = "-".join(s[:3])
                _a = "-".join(s[3:])

                _d = {"value": _a, "marked": False, "poses": [(c, r)]}
                d = cache.setdefault(_b, _d)
                if d != _d:
                    if d["value"] == _a:
                        if not d["marked"]:
                            d["poses"].append((c, r))

                        else:
                            cell.fill = mark_fl

                    else:
                        d["marked"] = True
                        cell.fill = mark_fl
                        for c, r in d["poses"]:
                            sheet.cell(c, r).fill = mark_fl

                        d["poses"] = []

                if not v.startswith("#"):
                    plates.add(DummyCell(cell))

        plates = list(plates)
        plates.sort()

        for i, p in enumerate(plates):
            _cell = p.cell
            cell = sheet2.cell(i + 1 + 1, 1)
            cell.value = _cell.value
            if _cell.fill == mark_fl:
                cell.fill = mark_fl

        wb.save(f"xlxs_folder/out/search/{col}.xlsx")
